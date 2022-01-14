'''
Mora-DSE BLURSONG 2021
A greedy and singel model version

HW variable params:
------------------------------------------------------------------------------------------------------------------------
HW       |  latancy      area      power     energy    |   
MNSIM    |  ns           um2       W         nJ        |   ①Tile BW = GB/s      ②Tile nums(1d, of a chip)
maestro  |  cycles/ns    um2       uW        nJ        |   ①NoC BW  = kB/s      ②PE nums                     L2 = Byte
mora     |  ns           um2       W         nJ        |     Top BW = GB/s                                     L2 = MB
------------------------------------------------------------------------------------------------------------------------
Scenarios:(modified)
-------------------------------------------------------------------------
Scenario     |  PEs       Tiles                  BW/GB/s       L2(GLB)/MB
edge         |  1024      24 * 24 - 4  chip      16            6
desktop      |  4096      48 * 48 - 16 chip      64            10
cloud        |  16384     96 * 96 - 64 chip      256           20
-------------------------------------------------------------------------

'''
import copy
import os
from random import choice
import sys
import argparse
from typing import Container
import re
import numpy as np
import pandas as pd
import subprocess as SP
import multiprocessing as MP
import torch
import MNSIM
import mora
import openpyxl
import math


def set_path(model, dataflow):
    global home_path, hw_config_path
    home_path = os.path.dirname(__file__)
    hw_config_path = os.path.abspath(os.path.join(home_path, "hw_config.m"))
    MNSIM_path = os.path.abspath(os.path.join(home_path, "MNSIM"))
    maestro_path = os.path.abspath(os.path.join(home_path, "maestro"))
    model_path = os.path.abspath(os.path.join(home_path, 'model/' + model))
    output_path = os.path.join(home_path, 'output/' + model)
    if os.path.exists(output_path):
        SP.run('rm *.csv', cwd=output_path, shell=True)
    if os.path.exists(model_path):
        SP.run('rm ' + model + '.csv', cwd=model_path, shell=True)
        SP.run('rm ' + model + '_dla_' + dataflow + '.m', cwd=model_path, shell=True)
        SP.run('rm ' + model + '_dla_model.m', cwd=model_path, shell=True)
    sys.path.append(home_path)
    sys.path.append(MNSIM_path)
    sys.path.append(maestro_path)


def set_parser():
    parser = argparse.ArgumentParser(description='mora dse parser')
    parser.add_argument('--dataflow', type=str, default='kcp_ws', choices=['ykp_os', 'yxp_os', 'kcp_ws', 'xp_ws', 'rs'])
    parser.add_argument('--model',
                        type=str,
                        default='vgg16',
                        choices=['alexnet', 'vgg16', 'vgg19', 'resnet18', 'resnet34', 'resnet50', 'resnext50', 'mobilenet_v2', 'shufflenet_v2', 'unet'])
    parser.add_argument('--scenario', type=str, default='edge', choices=['edge', 'desktop', 'cloud'])
    return parser


def hw_init_depr(hw_config_path):
    hw_dicts = {}
    with open(hw_config_path, 'r') as fs:
        for lines in fs:
            try:
                params, values = lines.split(':')
                hw_dicts[params.strip()] = int(values.strip())
            except ValueError:
                print("check your config file.")
    return hw_dicts


def hw_init(model, max_hw_param_dicts):
    hw_dicts = {}
    modeldata = openpyxl.load_workbook('model/modelmem.xlsx')
    sheet1 = modeldata.worksheets[0]
    for rowidx in range(1, sheet1.max_row):
        if sheet1.cell(rowidx, 1).value == model:
            xbarnum = sheet1.cell(rowidx, 5).value
            addtiles = sheet1.cell(rowidx, 9).value
            break
    rrampes = math.ceil(xbarnum / 8.0)
    rramtiles = math.ceil(rrampes / 16.0)
    hw_dicts['tiles-buildin'] = math.ceil(math.sqrt(rramtiles))
    assert hw_dicts['tiles-buildin'] <= max_hw_param_dicts['tiles'], "[mora][HW] scenario too small for RRAM."
    # change buildin strategy
    hw_dicts['tiles-buildin'] = hw_dicts['tiles-buildin'] + addtiles
    hw_dicts['tiles'] = int(hw_dicts['tiles-buildin'] / 4)
    hw_dicts['pes'] = int(max_hw_param_dicts['pes'] / 4) - int(max_hw_param_dicts['pes'] / 16)
    hw_dicts['glb_size'] = int(max_hw_param_dicts['glb_size'])
    hw_dicts['dla_bw'] = int(max_hw_param_dicts['bw'] / 4)
    hw_dicts['rram_bw'] = int(max_hw_param_dicts['bw'] / 4)
    return hw_dicts


def set_hw_range(scenario):
    if scenario == 'edge':
        mpes = 1024
        mtiles = 24
        mglb_size = 6  # MB
        mbw = 16  # GB/s
    elif scenario == 'desktop':
        mpes = 4096
        mtiles = 48
        mglb_size = 10
        mbw = 64
    elif scenario == 'cloud':
        mpes = 16384
        mtiles = 96
        mglb_size = 20
        mbw = 256
    max_hw_param_dicts = {}
    max_hw_param_dicts['pes'] = mpes
    max_hw_param_dicts['tiles'] = mtiles
    max_hw_param_dicts['glb_size'] = mglb_size
    max_hw_param_dicts['bw'] = mbw
    max_hw_param_dicts['tiles-buildin'] = mtiles
    return max_hw_param_dicts


if __name__ == "__main__":
    parser = set_parser()
    args = parser.parse_args()
    set_path(args.model, args.dataflow)
    max_hw_param_dicts = set_hw_range(args.scenario)
    ini_hw_param_dicts = hw_init(args.model, max_hw_param_dicts)
    max_hw_param_dicts['tiles-buildin'] = ini_hw_param_dicts['tiles-buildin']
    max_hw_param_dicts['tiles'] = ini_hw_param_dicts['tiles-buildin']

    dla = mora.HW.DLA(max_hw_param_dicts, args.dataflow, home_path)
    rram = mora.HW.RRAM(max_hw_param_dicts, home_path)
    mora.api.check_mora_csv(home_path, args.model)
    mora.api.remove_csv_bn(home_path, args.model)
    mora.api.gemm(home_path, args.model, args.dataflow)
    # TODO: new workload

    print("[mora] Init indicator.")
    dla.invoke_maestro(args.model)
    dla.export(args.model)
    rram.invoke_MNSIM(args.model, args.dataflow)
    edp_cons = mora.api.EDP(args.model, args.dataflow, home_path)
    area_cons = mora.api.area(args.model, args.dataflow, home_path)
    mora.schedule.greedy_schedule(DLA=dla,
                                  RRAM=rram,
                                  model=args.model,
                                  EDP_cons=edp_cons,
                                  area_cons=area_cons,
                                  ini_hw_param_dicts=ini_hw_param_dicts,
                                  max_param_dicts=max_hw_param_dicts,
                                  scenario=args.scenario)
    # TODO: new schedule

    mora.schedule.fifo_schedule(DLA=dla,
                                RRAM=rram,
                                model=args.model,
                                EDP_cons=edp_cons,
                                area_cons=area_cons,
                                ini_hw_param_dicts=ini_hw_param_dicts,
                                max_param_dicts=max_hw_param_dicts)
